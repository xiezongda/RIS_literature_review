from __future__ import annotations

import hashlib
import json
import os
import re
import time
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

import requests

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from literature_annotations import LITERATURE_ANNOTATIONS
except ImportError:
    LITERATURE_ANNOTATIONS = {}


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
EXCEL_DIR = ROOT / "output" / "excel"
MARKDOWN_DIR = ROOT / "output" / "markdown"
CACHE_DIR = ROOT / "output" / "cache"
CONFIG_DIR = ROOT / "config"
EXCEL_PATH = EXCEL_DIR / "literature_records.xlsx"
CARDS_PATH = MARKDOWN_DIR / "literature_cards.md"
SUMMARY_PATH = MARKDOWN_DIR / "literature_summary.md"
AI_CACHE_PATH = CACHE_DIR / "literature_ai_annotations.json"
LLM_CONFIG_PATH = CONFIG_DIR / "deepseek_v4pro.json"
REVIEW_TEMPLATE_PATH = ROOT / "review_templete.md"
SUMMARY_TEMPLATE_PATH = ROOT / "sunmary_templete.md"
PROMPTS_DIR = ROOT / "propmts"
LEGACY_REVIEW_TEMPLATE_PATH = PROMPTS_DIR / "review_templete.md"
LEGACY_SUMMARY_TEMPLATE_PATH = PROMPTS_DIR / "sunmary_templete.md"
AI_ANALYSIS_KEYS = (
    "broad_direction",
    "medium_direction",
    "abstract_summary",
    "study_object",
    "methods",
    "core_result",
    "fcva_connection",
    "complexity",
    "review_sentence",
)
RUNTIME_ANNOTATIONS: dict[int, dict[str, str]] = {}

TITLE_TAGS = ("TI", "T1", "CT", "BT")
AUTHOR_TAGS = ("AU", "A1", "A2", "A3", "A4")
YEAR_TAGS = ("PY", "Y1", "Y2", "DA")
JOURNAL_TAGS = ("JO", "JF", "JA", "J1", "J2", "T2")
DOI_TAGS = ("DO",)
KEYWORD_TAGS = ("KW", "DE")
ABSTRACT_TAGS = ("AB", "N2")

RIS_LINE_RE = re.compile(r"^([A-Z0-9]{2})\s{2}-\s?(.*)$")
ENDNOTE_LINE_RE = re.compile(r"^%([A-Z0-9])\s*(.*)$")
YEAR_RE = re.compile(r"(19|20)\d{2}")
DOI_RE = re.compile(r"10\.\d{4,9}/[-._;()/:A-Z0-9]+", re.IGNORECASE)
SENTENCE_SPLIT_RE = re.compile(r"(?<=[。！？.!?])\s+|\n+")
ENDNOTE_TAG_MAP = {
    "0": "TY",
    "A": "AU",
    "B": "T2",
    "D": "PY",
    "J": "JO",
    "K": "KW",
    "N": "IS",
    "O": "N1",
    "P": "SP",
    "R": "DO",
    "T": "TI",
    "U": "UR",
    "V": "VL",
    "X": "AB",
    "Z": "N1",
}


@dataclass
class LiteratureRecord:
    title: str = ""
    authors: list[str] = field(default_factory=list)
    year: str = ""
    journal: str = ""
    doi: str = ""
    keywords: list[str] = field(default_factory=list)
    abstract: str = ""
    source_files: list[str] = field(default_factory=list)
    duplicate_count: int = 1
    broad_direction: str = ""
    medium_direction: str = ""


def main() -> None:
    global RUNTIME_ANNOTATIONS

    ensure_directories()
    load_env_file(ROOT / ".env")
    ris_files = sorted(RAW_DIR.rglob("*.ris"))

    raw_records: list[LiteratureRecord] = []
    for ris_file in ris_files:
        raw_records.extend(parse_ris_file(ris_file))

    records = deduplicate_records(raw_records)
    RUNTIME_ANNOTATIONS = prepare_runtime_annotations(records)

    for index, record in enumerate(records, start=1):
        analysis = build_analysis(index, record)
        record.broad_direction = analysis["broad_direction"]
        record.medium_direction = analysis["medium_direction"]

    write_excel(records, EXCEL_PATH)
    write_literature_cards(records, CARDS_PATH)
    write_summary(records, SUMMARY_PATH)

    print(f"RIS 文件数量：{len(ris_files)}")
    print(f"原始文献条目：{len(raw_records)}")
    print(f"去重后条目：{len(records)}")
    print(f"AI/缓存标注条目：{len(RUNTIME_ANNOTATIONS)}")
    print(f"Excel 输出：{EXCEL_PATH}")
    print(f"文献卡片：{CARDS_PATH}")
    print(f"分类总结：{SUMMARY_PATH}")


def ensure_directories() -> None:
    for path in (RAW_DIR, EXCEL_DIR, MARKDOWN_DIR, CACHE_DIR, CONFIG_DIR):
        path.mkdir(parents=True, exist_ok=True)


def load_env_file(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def prepare_runtime_annotations(records: list[LiteratureRecord]) -> dict[int, dict[str, str]]:
    annotations: dict[int, dict[str, str]] = {}
    cache = load_annotation_cache()
    cache_records = cache.setdefault("records", {})
    refresh_ai = os.getenv("LITERATURE_REFRESH_AI", "").strip() == "1"
    use_ai = os.getenv("LITERATURE_USE_AI", "1").strip() != "0"
    llm_config = load_llm_config()
    api_key = str(llm_config.get("api_key", "")).strip()
    cache_changed = False
    missing: list[tuple[int, LiteratureRecord, str]] = []

    for index, record in enumerate(records, start=1):
        key = record_cache_key(record)
        cached = cache_records.get(key, {}) if not refresh_ai else {}
        cached_analysis = normalize_analysis(cached.get("analysis")) if cached else {}
        if cached_analysis:
            annotations[index] = cached_analysis
            continue

        preset = preset_annotation_for_current_records(index, records)
        if preset and not refresh_ai:
            annotations[index] = preset
            cache_records[key] = cache_entry(record, preset, source="preset")
            cache_changed = True
            continue

        missing.append((index, record, key))

    if missing and api_key and use_ai:
        print(f"需要 AI 阅读的新文献：{len(missing)} 条")
        for position, (index, record, key) in enumerate(missing, start=1):
            print(f"AI 阅读 {position}/{len(missing)}：{record.title[:80] or '未命名文献'}")
            try:
                analysis = annotate_record_with_llm(index, record, llm_config)
            except Exception as exc:
                print(f"AI 阅读失败，改用规则兜底：第 {index} 条，{exc}")
                analysis = {}

            if analysis:
                annotations[index] = analysis
                cache_records[key] = cache_entry(record, analysis, source=str(llm_config.get("provider", "llm")))
                cache_changed = True
                save_annotation_cache(cache)
                time.sleep(float(llm_config.get("sleep_seconds", 0.2)))
    elif missing:
        if api_key and not use_ai:
            print("已设置 LITERATURE_USE_AI=0，未调用 AI；缺失条目会用规则兜底。")
        else:
            print(f"未在 {LLM_CONFIG_PATH} 或环境变量中检测到 API key；新文献无法自动 AI 阅读，缺失条目会用规则兜底。")

    if cache_changed:
        save_annotation_cache(cache)

    return annotations


def load_annotation_cache() -> dict[str, Any]:
    if not AI_CACHE_PATH.exists():
        return {"version": 1, "records": {}}
    try:
        return json.loads(AI_CACHE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"version": 1, "records": {}}


def save_annotation_cache(cache: dict[str, Any]) -> None:
    AI_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    AI_CACHE_PATH.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_llm_config() -> dict[str, Any]:
    config: dict[str, Any] = {
        "provider": "deepseek",
        "api_key": "",
        "base_url": "https://api.deepseek.com",
        "model": "deepseek-v4-pro",
        "endpoint": "chat_completions",
        "temperature": 0.2,
        "max_tokens": 2200,
        "timeout_seconds": 180,
        "sleep_seconds": 0.2,
        "json_mode": True,
        "thinking": {"type": "enabled"},
        "reasoning_effort": "high",
    }

    if LLM_CONFIG_PATH.exists():
        try:
            file_config = json.loads(LLM_CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(file_config, dict):
                config.update(file_config)
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"DeepSeek 配置文件不是合法 JSON：{LLM_CONFIG_PATH}，{exc}") from exc

    # Environment variables are convenient for temporary overrides.
    env_api_key = os.getenv("DEEPSEEK_API_KEY") or os.getenv("OPENAI_API_KEY") or os.getenv("LITERATURE_API_KEY")
    if env_api_key:
        config["api_key"] = env_api_key
    if os.getenv("DEEPSEEK_BASE_URL"):
        config["base_url"] = os.getenv("DEEPSEEK_BASE_URL")
    if os.getenv("DEEPSEEK_MODEL"):
        config["model"] = os.getenv("DEEPSEEK_MODEL")

    placeholder_keys = {"", "填入你的 DeepSeek API Key", "sk-your-deepseek-api-key"}
    if str(config.get("api_key", "")).strip() in placeholder_keys:
        config["api_key"] = ""
    return config


def record_cache_key(record: LiteratureRecord) -> str:
    material = "\n".join(
        [
            normalize_doi(record.doi),
            normalize_title(record.title),
            record.year,
            record.journal.casefold(),
            clean_spaces(record.abstract),
        ]
    )
    return hashlib.sha256(material.encode("utf-8")).hexdigest()


def cache_entry(record: LiteratureRecord, analysis: dict[str, str], source: str) -> dict[str, Any]:
    return {
        "source": source,
        "model": "",
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "title": record.title,
        "doi": record.doi,
        "analysis": analysis,
    }


def preset_annotation_for_current_records(index: int, records: list[LiteratureRecord]) -> dict[str, str]:
    if not LITERATURE_ANNOTATIONS:
        return {}
    if len(records) != len(LITERATURE_ANNOTATIONS):
        return {}
    source_names = {name for record in records for name in record.source_files}
    if source_names != {"YSZ_calculation_computation.ris"}:
        return {}
    return normalize_analysis(LITERATURE_ANNOTATIONS.get(index))


def normalize_analysis(value: Any) -> dict[str, str]:
    if not isinstance(value, dict):
        return {}
    normalized: dict[str, str] = {}
    for key in AI_ANALYSIS_KEYS:
        text = clean_spaces(str(value.get(key, "")))
        if text:
            normalized[key] = text
    if all(key in normalized for key in AI_ANALYSIS_KEYS):
        return normalized
    return {}


def annotate_record_with_llm(index: int, record: LiteratureRecord, config: dict[str, Any]) -> dict[str, str]:
    endpoint_type = str(config.get("endpoint", "chat_completions"))
    if endpoint_type == "responses":
        return annotate_record_with_responses_api(index, record, config)
    return annotate_record_with_chat_completions(index, record, config)


def annotate_record_with_chat_completions(index: int, record: LiteratureRecord, config: dict[str, Any]) -> dict[str, str]:
    endpoint = str(config.get("base_url", "https://api.deepseek.com")).rstrip("/") + "/chat/completions"
    payload: dict[str, Any] = {
        "model": config.get("model", "deepseek-v4-pro"),
        "messages": [
            {"role": "system", "content": build_ai_instructions()},
            {"role": "user", "content": build_ai_prompt(index, record)},
        ],
        "temperature": float(config.get("temperature", 0.2)),
        "max_tokens": int(config.get("max_tokens", 2200)),
    }
    if config.get("json_mode", True):
        payload["response_format"] = {"type": "json_object"}
    if config.get("thinking"):
        payload["thinking"] = config["thinking"]
    if config.get("reasoning_effort"):
        payload["reasoning_effort"] = config["reasoning_effort"]

    try:
        data = post_llm_json(endpoint, payload, config)
    except RuntimeError as exc:
        if "response_format" not in str(exc):
            raise
        payload.pop("response_format", None)
        data = post_llm_json(endpoint, payload, config)
    content = data["choices"][0]["message"].get("content", "")
    analysis = normalize_analysis(json.loads(extract_json_object(content)))
    if not analysis:
        raise RuntimeError("模型返回的 JSON 字段不完整")
    return analysis


def annotate_record_with_responses_api(index: int, record: LiteratureRecord, config: dict[str, Any]) -> dict[str, str]:
    endpoint = str(config.get("base_url", "https://api.openai.com/v1")).rstrip("/") + "/responses"
    payload = {
        "model": config.get("model", "gpt-4o-mini"),
        "instructions": build_ai_instructions(),
        "input": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": build_ai_prompt(index, record),
                    }
                ],
            }
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "literature_annotation",
                "strict": True,
                "schema": analysis_json_schema(),
            }
        },
        "temperature": float(config.get("temperature", 0.2)),
        "max_output_tokens": int(config.get("max_tokens", 2200)),
    }
    data = post_llm_json(endpoint, payload, config)
    analysis = normalize_analysis(json.loads(extract_response_text(data)))
    if not analysis:
        raise RuntimeError("模型返回的 JSON 字段不完整")
    return analysis


def post_llm_json(endpoint: str, payload: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    headers = {
        "Authorization": f"Bearer {config['api_key']}",
        "Content-Type": "application/json",
    }
    last_error: Exception | None = None
    for attempt in range(1, 4):
        try:
            response = requests.post(
                endpoint,
                headers=headers,
                json=payload,
                timeout=int(config.get("timeout_seconds", 180)),
            )
            if response.status_code >= 400:
                raise RuntimeError(f"HTTP {response.status_code}: {response.text[:500]}")
            return response.json()
        except Exception as exc:
            last_error = exc
            if attempt < 3:
                time.sleep(2 * attempt)
    raise RuntimeError(str(last_error))


def build_ai_instructions() -> str:
    review_prompt = read_template_section(
        REVIEW_TEMPLATE_PATH,
        "AI_READING_PROMPT",
        default_review_ai_prompt(),
    )
    summary_prompt = read_template_section(
        SUMMARY_TEMPLATE_PATH,
        "SUMMARY_PROMPT",
        default_summary_prompt(),
    )
    return "\n\n".join(
        [
            review_prompt,
            "总结文件分类要求如下，生成 broad_direction 和 medium_direction 时必须服务于该分类逻辑：",
            summary_prompt,
            "输出必须是一个合法 json 对象，不要输出 Markdown 或解释文字。",
        ]
    )


def build_ai_prompt(index: int, record: LiteratureRecord) -> str:
    return "\n".join(
        [
            f"文献序号：{index}",
            f"标题：{record.title or '未提供'}",
            f"作者：{'; '.join(record.authors) or '未提供'}",
            f"年份：{record.year or '未提供'}",
            f"期刊：{record.journal or '未提供'}",
            f"DOI：{record.doi or '未提供'}",
            f"关键词：{'; '.join(record.keywords) or '未提供'}",
            f"摘要：{truncate(record.abstract, 7000) or '未提供'}",
            "",
            "文献卡片最终会按 review_templete.md 中的 CARD_TEMPLATE 渲染，请生成能填入该模板的内容。",
            "",
            "请输出以下字段：",
            "broad_direction：研究大方向，适合综述目录一级分类。",
            "medium_direction：中等方向，适合综述目录二级分类。",
            "abstract_summary：1 句话概括摘要。",
            "study_object：研究对象。",
            "methods：研究方法。",
            "core_result：核心结果。",
            "fcva_connection：和“常温 FCVA 制备 YSZ 薄膜”可结合的点。",
            "complexity：实验复杂性和可实现性。",
            "review_sentence：可直接放入论文综述的一句话。",
        ]
    )


def analysis_json_schema() -> dict[str, Any]:
    properties = {
        key: {
            "type": "string",
            "minLength": 1,
        }
        for key in AI_ANALYSIS_KEYS
    }
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": properties,
        "required": list(AI_ANALYSIS_KEYS),
    }


def extract_response_text(data: dict[str, Any]) -> str:
    if isinstance(data.get("output_text"), str):
        return data["output_text"]

    texts: list[str] = []
    for item in data.get("output", []):
        for content in item.get("content", []):
            text = content.get("text")
            if isinstance(text, str):
                texts.append(text)
    if texts:
        return "\n".join(texts)
    raise RuntimeError("无法从 API 响应中提取文本")


def extract_json_object(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)
    if text.startswith("{") and text.endswith("}"):
        return text

    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        return text[start : end + 1]
    raise RuntimeError("模型响应中没有找到 JSON 对象")


def read_template_section(path: Path, section: str, default: str) -> str:
    if not path.exists():
        if path == REVIEW_TEMPLATE_PATH:
            path = LEGACY_REVIEW_TEMPLATE_PATH
        elif path == SUMMARY_TEMPLATE_PATH:
            path = LEGACY_SUMMARY_TEMPLATE_PATH

    if not path.exists():
        return default

    text = path.read_text(encoding="utf-8")
    start_marker = f"<!-- {section}_START -->"
    end_marker = f"<!-- {section}_END -->"
    start = text.find(start_marker)
    end = text.find(end_marker)
    if start == -1 or end == -1 or end <= start:
        return default
    content = text[start + len(start_marker) : end].strip()
    return content or default


def render_template(template: str, context: dict[str, Any]) -> str:
    rendered = template
    for key, value in context.items():
        rendered = rendered.replace("{{" + key + "}}", str(value))
    return rendered


def default_review_ai_prompt() -> str:
    return (
        "你是材料科学与薄膜制备方向的文献综述助手。"
        "请阅读用户提供的 RIS 题录字段，重点依据 title、abstract、keywords、journal 和 doi，"
        "为主题“常温 FCVA 制备 YSZ 薄膜”生成中文文献卡片分析。"
        "不要编造题录中没有的信息；如果摘要为空，必须明确写“题录缺摘要，需阅读全文验证”。"
    )


def default_summary_prompt() -> str:
    return (
        "请将文献按照研究大方向到中等方向分类；分类应适合 Obsidian 总览，"
        "并能通过 Dataview 表格链接到 literature_cards.md 中对应文献卡片。"
    )


def parse_ris_file(path: Path) -> list[LiteratureRecord]:
    text = read_text_with_fallback(path)
    entries = parse_ris_entries(text)
    if not entries and looks_like_endnote_export(text):
        entries = parse_endnote_entries(text)
    return [entry_to_record(entry, path.name) for entry in entries]


def read_text_with_fallback(path: Path) -> str:
    encodings = ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1")
    last_error: UnicodeDecodeError | None = None
    for encoding in encodings:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return path.read_text()


def parse_ris_entries(text: str) -> list[dict[str, list[str]]]:
    entries: list[dict[str, list[str]]] = []
    current: dict[str, list[str]] = defaultdict(list)
    current_tag: str | None = None

    for raw_line in text.splitlines():
        line = raw_line.rstrip("\n\r")
        match = RIS_LINE_RE.match(line)
        if match:
            tag, value = match.group(1), clean_spaces(match.group(2))
            if tag == "ER":
                if current:
                    entries.append(dict(current))
                current = defaultdict(list)
                current_tag = None
                continue
            current[tag].append(value)
            current_tag = tag
        elif current_tag and line.strip():
            current[current_tag][-1] = clean_spaces(f"{current[current_tag][-1]} {line.strip()}")

    if current:
        entries.append(dict(current))

    return entries


def looks_like_endnote_export(text: str) -> bool:
    return bool(re.search(r"(?m)^%0\s+", text))


def parse_endnote_entries(text: str) -> list[dict[str, list[str]]]:
    entries: list[dict[str, list[str]]] = []
    current: dict[str, list[str]] = defaultdict(list)
    current_tag: str | None = None

    for raw_line in text.splitlines():
        line = raw_line.rstrip("\n\r")
        match = ENDNOTE_LINE_RE.match(line)
        if match:
            raw_tag, value = match.group(1), clean_spaces(match.group(2))
            tag = ENDNOTE_TAG_MAP.get(raw_tag, f"EN_{raw_tag}")

            if raw_tag == "0" and current:
                entries.append(dict(current))
                current = defaultdict(list)

            current[tag].append(value)
            current_tag = tag
        elif current_tag and line.strip():
            current[current_tag][-1] = clean_spaces(f"{current[current_tag][-1]} {line.strip()}")

    if current:
        entries.append(dict(current))

    return entries


def entry_to_record(entry: dict[str, list[str]], source_file: str) -> LiteratureRecord:
    title = first_value(entry, TITLE_TAGS)
    authors = unique_non_empty(values_for(entry, AUTHOR_TAGS))
    year = extract_year(first_value(entry, YEAR_TAGS))
    journal = first_value(entry, JOURNAL_TAGS)
    doi = normalize_doi(first_value(entry, DOI_TAGS) or find_doi_in_entry(entry))
    keywords = split_keywords(values_for(entry, KEYWORD_TAGS))
    abstract = join_unique(values_for(entry, ABSTRACT_TAGS))

    return LiteratureRecord(
        title=title,
        authors=authors,
        year=year,
        journal=journal,
        doi=doi,
        keywords=keywords,
        abstract=abstract,
        source_files=[source_file],
    )


def values_for(entry: dict[str, list[str]], tags: Iterable[str]) -> list[str]:
    values: list[str] = []
    for tag in tags:
        values.extend(entry.get(tag, []))
    return [clean_spaces(value) for value in values if clean_spaces(value)]


def first_value(entry: dict[str, list[str]], tags: Iterable[str]) -> str:
    values = values_for(entry, tags)
    return values[0] if values else ""


def clean_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def join_unique(values: Iterable[str], separator: str = " ") -> str:
    return separator.join(unique_non_empty(values)).strip()


def unique_non_empty(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = clean_spaces(value)
        key = cleaned.casefold()
        if cleaned and key not in seen:
            seen.add(key)
            result.append(cleaned)
    return result


def split_keywords(values: Iterable[str]) -> list[str]:
    parts: list[str] = []
    for value in values:
        parts.extend(re.split(r"\s*[;,；]\s*|\s{2,}", value))
    return unique_non_empty(parts)


def extract_year(value: str) -> str:
    match = YEAR_RE.search(value or "")
    return match.group(0) if match else ""


def find_doi_in_entry(entry: dict[str, list[str]]) -> str:
    for values in entry.values():
        for value in values:
            match = DOI_RE.search(value)
            if match:
                return match.group(0)
    return ""


def normalize_doi(value: str) -> str:
    if not value:
        return ""
    value = value.strip()
    match = DOI_RE.search(value)
    if match:
        value = match.group(0)
    value = re.sub(r"^https?://(dx\.)?doi\.org/", "", value, flags=re.IGNORECASE)
    value = re.sub(r"^doi:\s*", "", value, flags=re.IGNORECASE)
    return value.strip(" .;,").lower()


def deduplicate_records(records: list[LiteratureRecord]) -> list[LiteratureRecord]:
    kept: list[LiteratureRecord] = []
    doi_index: dict[str, int] = {}

    for record in records:
        duplicate_index: int | None = None

        if record.doi and record.doi in doi_index:
            duplicate_index = doi_index[record.doi]
        else:
            duplicate_index = find_title_duplicate(record, kept)

        if duplicate_index is None:
            kept.append(record)
            if record.doi:
                doi_index[record.doi] = len(kept) - 1
        else:
            merge_records(kept[duplicate_index], record)
            if kept[duplicate_index].doi:
                doi_index[kept[duplicate_index].doi] = duplicate_index

    return kept


def find_title_duplicate(record: LiteratureRecord, kept: list[LiteratureRecord]) -> int | None:
    title_key = normalize_title(record.title)
    if not title_key:
        return None

    best_index: int | None = None
    best_score = 0.0
    for index, existing in enumerate(kept):
        if record.doi and existing.doi and record.doi != existing.doi:
            continue
        existing_key = normalize_title(existing.title)
        if not existing_key:
            continue

        if title_key == existing_key:
            score = 1.0
        else:
            score = SequenceMatcher(None, title_key, existing_key).ratio()

        if score > best_score:
            best_score = score
            best_index = index

    if best_score >= 0.92:
        return best_index
    return None


def normalize_title(title: str) -> str:
    title = title.casefold()
    title = re.sub(r"[^\w\u4e00-\u9fff]+", " ", title)
    return clean_spaces(title)


def merge_records(target: LiteratureRecord, incoming: LiteratureRecord) -> None:
    target.title = choose_richer_text(target.title, incoming.title)
    target.authors = unique_non_empty([*target.authors, *incoming.authors])
    target.year = target.year or incoming.year
    target.journal = target.journal or incoming.journal
    target.doi = target.doi or incoming.doi
    target.keywords = unique_non_empty([*target.keywords, *incoming.keywords])
    target.abstract = choose_richer_text(target.abstract, incoming.abstract)
    target.source_files = unique_non_empty([*target.source_files, *incoming.source_files])
    target.duplicate_count += incoming.duplicate_count


def choose_richer_text(current: str, incoming: str) -> str:
    if not current:
        return incoming
    if not incoming:
        return current
    return incoming if len(incoming) > len(current) else current


def classify_record(record: LiteratureRecord) -> tuple[str, str]:
    text = searchable_text(record)

    if has_any(text, "review", "systematic review", "meta-analysis", "overview", "综述"):
        return "综述、理论与方法框架", "综述研究"
    if has_any(text, "simulation", "model", "finite element", "density functional", "dft", "phase-field", "模拟", "建模"):
        return "综述、理论与方法框架", "建模与模拟"

    if has_any(text, "fcva", "filtered cathodic vacuum arc", "cathodic vacuum arc", "vacuum arc", "阴极真空弧", "真空弧"):
        return "薄膜与涂层制备", "FCVA/阴极真空弧沉积"
    if has_any(text, "sputter", "pvd", "pld", "physical vapor", "magnetron", "evaporation", "溅射", "物理气相", "脉冲激光"):
        return "薄膜与涂层制备", "PVD/溅射/PLD 制备"
    if has_any(text, "chemical vapor", "cvd", "ald", "atomic layer", "化学气相", "原子层沉积"):
        return "薄膜与涂层制备", "CVD/ALD 制备"
    if has_any(text, "sol-gel", "sol gel", "hydrothermal", "precipitation", "wet chemical", "溶胶", "水热", "湿化学"):
        return "薄膜与涂层制备", "湿化学与溶胶-凝胶"

    if has_any(text, "ysz", "yttria-stabilized zirconia", "yttria stabilized zirconia", "zirconia", "zro2", "氧化锆", "钇稳定"):
        if has_any(text, "thermal barrier", "tbc", "热障"):
            return "YSZ/氧化锆材料性能", "热障涂层与热稳定性"
        if has_any(text, "ionic conductivity", "electrolyte", "sofc", "fuel cell", "离子电导", "电解质", "燃料电池"):
            return "YSZ/氧化锆材料性能", "离子导电与电化学性能"
        return "YSZ/氧化锆材料性能", "相结构、稳定化与基础性能"

    if has_any(text, "xrd", "sem", "tem", "xps", "raman", "afm", "microstructure", "表征", "相结构", "显微"):
        return "表征与性能评价", "微结构与相组成表征"
    if has_any(text, "hardness", "wear", "friction", "adhesion", "mechanical", "硬度", "磨损", "摩擦", "结合力"):
        return "表征与性能评价", "力学与摩擦磨损性能"
    if has_any(text, "optical", "electrical", "thermal conductivity", "dielectric", "光学", "电学", "热导"):
        return "表征与性能评价", "光学/电学/热学性能"

    if has_any(text, "sensor", "biomedical", "implant", "catalyst", "membrane", "传感", "生物", "催化", "膜"):
        return "应用与器件", "功能器件与应用"

    return "其他与待人工复核", "待人工分类"


def searchable_text(record: LiteratureRecord) -> str:
    return " ".join(
        [
            record.title,
            " ".join(record.keywords),
            record.abstract,
            record.journal,
        ]
    ).casefold()


def has_any(text: str, *needles: str) -> bool:
    return any(needle.casefold() in text for needle in needles)


def write_excel(records: list[LiteratureRecord], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "literature_records"

    headers = [
        "index",
        "title",
        "authors",
        "year",
        "journal",
        "doi",
        "keywords",
        "abstract",
        "abstract_summary",
        "study_object",
        "methods",
        "core_result",
        "fcva_connection",
        "complexity",
        "review_sentence",
        "broad_direction",
        "medium_direction",
        "source_files",
        "duplicate_count",
    ]
    sheet.append(headers)

    for index, record in enumerate(records, start=1):
        analysis = build_analysis(index, record)
        sheet.append(
            [
                index,
                record.title,
                "; ".join(record.authors),
                record.year,
                record.journal,
                record.doi,
                "; ".join(record.keywords),
                record.abstract,
                analysis["abstract_summary"],
                analysis["study_object"],
                analysis["methods"],
                analysis["core_result"],
                analysis["fcva_connection"],
                analysis["complexity"],
                analysis["review_sentence"],
                analysis["broad_direction"],
                analysis["medium_direction"],
                "; ".join(record.source_files),
                record.duplicate_count,
            ]
        )

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 8,
        "B": 45,
        "C": 32,
        "D": 10,
        "E": 28,
        "F": 28,
        "G": 32,
        "H": 70,
        "I": 42,
        "J": 34,
        "K": 34,
        "L": 46,
        "M": 46,
        "N": 42,
        "O": 54,
        "P": 24,
        "Q": 26,
        "R": 24,
        "S": 15,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(records) + 1)}"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_literature_cards(records: list[LiteratureRecord], path: Path) -> None:
    lines: list[str] = [
        "# 文献卡片",
        "",
        f"> 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
    ]

    if not records:
        lines.extend(
            [
                "当前未在 `data/raw/` 中读取到 RIS 文献。把 EndNote 导出的 `.ris` 文件放入该目录后，重新运行 `python src/main.py`。",
                "",
            ]
        )
    else:
        for index, record in enumerate(records, start=1):
            lines.extend(render_record_card(index, record))
            lines.append("")

    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def render_record_card(index: int, record: LiteratureRecord) -> list[str]:
    title = record.title or "未命名文献"
    authors = format_authors(record.authors)
    keywords = "；".join(record.keywords) if record.keywords else "未提供"
    analysis = build_analysis(index, record)
    template = read_template_section(
        REVIEW_TEMPLATE_PATH,
        "CARD_TEMPLATE",
        default_card_template(),
    )
    context = {
        "index": index,
        "literature_id": f"L{index:03d}",
        "title": title,
        "title_inline": escape_inline_field(title),
        "year": record.year or "未提供",
        "authors": authors,
        "journal": record.journal or "未提供",
        "doi": record.doi or "未提供",
        "keywords": keywords,
        "abstract": record.abstract or "未提供",
        "source_files": "；".join(record.source_files),
        **analysis,
    }
    return (render_template(template, context).rstrip() + "\n").splitlines()


def default_card_template() -> str:
    return """## {{index}}. {{title}}

literature_id:: {{literature_id}}
title:: {{title_inline}}
year:: {{year}}
doi:: {{doi}}
broad_direction:: {{broad_direction}}
medium_direction:: {{medium_direction}}

- 年份：{{year}}
- 作者：{{authors}}
- 期刊：{{journal}}
- DOI：{{doi}}
- 关键词：{{keywords}}
- 摘要概括：{{abstract_summary}}
- 研究对象：{{study_object}}
- 研究方法：{{methods}}
- 核心结果：{{core_result}}
- “常温 FCVA 制备 YSZ 薄膜”可以结合的点：{{fcva_connection}}
- 实验复杂性和可实现性：{{complexity}}
- 可用于论文综述的句子：{{review_sentence}}
"""


def build_analysis(index: int, record: LiteratureRecord) -> dict[str, str]:
    fallback_study_object = infer_study_object(record)
    fallback_methods = infer_methods(record)
    fallback = {
        "broad_direction": record.broad_direction or classify_record(record)[0],
        "medium_direction": record.medium_direction or classify_record(record)[1],
        "abstract_summary": summarize_abstract(record.abstract),
        "study_object": fallback_study_object,
        "methods": fallback_methods,
        "core_result": infer_core_result(record),
        "fcva_connection": infer_fcva_connection(record),
        "complexity": infer_complexity(record),
        "review_sentence": build_review_sentence(record, fallback_study_object, fallback_methods),
    }

    override = RUNTIME_ANNOTATIONS.get(index, {})
    for key, value in override.items():
        if value:
            fallback[key] = clean_spaces(str(value))
    return fallback


def format_authors(authors: list[str], limit: int = 8) -> str:
    if not authors:
        return "未提供"
    shown = authors[:limit]
    suffix = " 等" if len(authors) > limit else ""
    return "；".join(shown) + suffix


def summarize_abstract(abstract: str, max_length: int = 360) -> str:
    if not abstract:
        return "原始记录未提供摘要，建议结合全文补充。"

    sentences = [sentence.strip() for sentence in SENTENCE_SPLIT_RE.split(abstract) if sentence.strip()]
    summary = " ".join(sentences[:2]) if sentences else abstract
    return truncate(summary, max_length)


def infer_study_object(record: LiteratureRecord) -> str:
    text = searchable_text(record)
    objects: list[str] = []
    object_terms = [
        ("YSZ/钇稳定氧化锆", ("ysz", "yttria-stabilized zirconia", "yttria stabilized zirconia", "钇稳定")),
        ("氧化锆/ZrO2", ("zirconia", "zro2", "氧化锆")),
        ("薄膜/涂层", ("thin film", "film", "coating", "薄膜", "涂层")),
        ("热障涂层", ("thermal barrier", "tbc", "热障")),
        ("固体氧化物燃料电池相关电解质", ("sofc", "fuel cell", "electrolyte", "燃料电池", "电解质")),
    ]
    for label, needles in object_terms:
        if any(needle in text for needle in needles):
            objects.append(label)

    if objects:
        return "、".join(unique_non_empty(objects))
    return "根据题录信息暂无法明确，建议阅读全文后补充。"


def infer_methods(record: LiteratureRecord) -> str:
    text = searchable_text(record)
    methods: list[str] = []
    method_terms = [
        ("FCVA/阴极真空弧沉积", ("fcva", "filtered cathodic vacuum arc", "cathodic vacuum arc", "vacuum arc", "阴极真空弧", "真空弧")),
        ("磁控溅射/PVD", ("sputter", "magnetron", "pvd", "physical vapor", "溅射", "物理气相")),
        ("PLD", ("pld", "pulsed laser", "脉冲激光")),
        ("CVD/ALD", ("cvd", "chemical vapor", "ald", "atomic layer", "化学气相", "原子层沉积")),
        ("溶胶-凝胶/湿化学", ("sol-gel", "sol gel", "hydrothermal", "wet chemical", "溶胶", "水热", "湿化学")),
        ("XRD", ("xrd", "x-ray diffraction", "x ray diffraction")),
        ("SEM/TEM", ("sem", "tem", "electron microscopy", "电子显微")),
        ("XPS/Raman/AFM", ("xps", "raman", "afm")),
        ("电化学测试", ("impedance", "eis", "electrochemical", "电化学", "阻抗")),
        ("模拟/建模", ("simulation", "model", "finite element", "dft", "模拟", "建模")),
    ]
    for label, needles in method_terms:
        if any(needle in text for needle in needles):
            methods.append(label)

    if methods:
        return "；".join(unique_non_empty(methods))
    return "题录中未显式给出方法，建议结合全文补充。"


def infer_core_result(record: LiteratureRecord) -> str:
    if not record.abstract:
        return "原始记录未提供摘要，核心结果需阅读全文后补充。"

    result_markers = (
        "show",
        "demonstrat",
        "result",
        "found",
        "indicat",
        "improv",
        "increase",
        "decrease",
        "enhanc",
        "表明",
        "结果",
        "发现",
        "提高",
        "降低",
        "增强",
    )
    sentences = [sentence.strip() for sentence in SENTENCE_SPLIT_RE.split(record.abstract) if sentence.strip()]
    for sentence in sentences:
        if has_any(sentence.casefold(), *result_markers):
            return truncate(sentence, 300)
    return truncate(sentences[-1] if sentences else record.abstract, 300)


def infer_fcva_connection(record: LiteratureRecord) -> str:
    text = searchable_text(record)
    if has_any(text, "fcva", "filtered cathodic vacuum arc", "cathodic vacuum arc", "vacuum arc", "阴极真空弧"):
        return "可直接对照 FCVA 工艺参数、离子能量、沉积温度、膜层致密性和缺陷控制。"
    if has_any(text, "ysz", "yttria", "zirconia", "zro2", "氧化锆", "钇稳定"):
        return "可为 YSZ 薄膜的相结构稳定、氧空位调控、热/电性能评价提供材料依据。"
    if has_any(text, "thin film", "film", "coating", "sputter", "pvd", "pld", "薄膜", "涂层", "溅射"):
        return "可作为室温薄膜沉积路线、表征指标和工艺-结构-性能关系的横向对比。"
    if has_any(text, "room temperature", "ambient temperature", "low temperature", "室温", "低温"):
        return "可用于论证低温或室温工艺的可行性、限制条件与后处理需求。"
    return "与主题的直接关联度需要人工复核，可优先检查材料体系、制备温度和表征指标是否可迁移。"


def infer_complexity(record: LiteratureRecord) -> str:
    text = searchable_text(record)
    if has_any(text, "fcva", "pld", "ald", "cvd", "magnetron", "sputter", "vacuum", "真空", "溅射", "原子层"):
        return "设备依赖较强，实验复杂性偏高；若已有真空沉积平台，可通过小样片和参数矩阵逐步验证。"
    if has_any(text, "sol-gel", "hydrothermal", "wet chemical", "溶胶", "水热", "湿化学"):
        return "工艺门槛中等，前驱体和热处理窗口需要控制；适合作为低成本对比路线。"
    if has_any(text, "simulation", "model", "review", "模拟", "建模", "综述"):
        return "实验负担较低，但需要转换为可验证的参数假设或评价指标。"
    return "复杂性暂无法从题录判断，建议优先核对设备条件、样品制备周期和关键表征需求。"


def build_review_sentence(record: LiteratureRecord, study_object: str, methods: str) -> str:
    authors = record.authors[0] if record.authors else "相关研究"
    year = f"（{record.year}）" if record.year else ""
    title_topic = record.title or study_object
    if methods.startswith("题录中未显式"):
        return f"{authors}{year}围绕“{truncate(title_topic, 80)}”开展研究，为理解 {study_object} 的结构与性能关系提供了参考。"
    return f"{authors}{year}以 {study_object} 为对象，采用 {methods} 等方法，讨论了“{truncate(title_topic, 80)}”相关问题，可为 YSZ 薄膜制备与性能评价提供参考。"


def write_summary(records: list[LiteratureRecord], path: Path) -> None:
    header_template = read_template_section(
        SUMMARY_TEMPLATE_PATH,
        "SUMMARY_HEADER_TEMPLATE",
        default_summary_header_template(),
    )
    lines: list[str] = render_template(
        header_template,
        {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "record_count": len(records),
            "cards_file": "literature_cards.md",
        },
    ).rstrip().splitlines()
    lines.append("")

    if not records:
        lines.extend(
            [
                "当前没有可分类文献。把 `.ris` 文件放入 `data/raw/` 后运行 `python src/main.py`。",
                "",
            ]
        )
    else:
        grouped: dict[str, dict[str, list[tuple[int, LiteratureRecord]]]] = defaultdict(lambda: defaultdict(list))
        for index, record in enumerate(records, start=1):
            grouped[record.broad_direction][record.medium_direction].append((index, record))

        for broad in sorted(grouped):
            lines.extend([f"## {broad}", ""])
            for medium in sorted(grouped[broad]):
                lines.extend([f"### {medium}", ""])
                lines.extend(render_dataview_table(grouped[broad][medium]))
                lines.append("")

    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def default_summary_header_template() -> str:
    return """# 文献分类总览

> 生成时间：{{generated_at}}

本文件按“研究大方向 → 中等方向”组织文献。每个分类下的表格由 Obsidian DataviewJS 构建，文献标题链接到 `{{cards_file}}` 中对应卡片。
"""


def render_dataview_table(items: list[tuple[int, LiteratureRecord]]) -> list[str]:
    rows = []
    for index, record in items:
        title = record.title or "未命名文献"
        heading = f"{index}. {title}"
        wikilink = f"[[literature_cards#{escape_wikilink_part(heading)}|{escape_wikilink_part(title)}]]"
        rows.append(
            {
                "link": wikilink,
                "year": record.year or "",
                "authors": format_authors(record.authors, limit=3),
                "journal": record.journal or "",
                "doi": record.doi or "",
            }
        )

    js_rows = ",\n  ".join(
        "{"
        f"link: {js_string(row['link'])}, "
        f"year: {js_string(row['year'])}, "
        f"authors: {js_string(row['authors'])}, "
        f"journal: {js_string(row['journal'])}, "
        f"doi: {js_string(row['doi'])}"
        "}"
        for row in rows
    )

    template = read_template_section(
        SUMMARY_TEMPLATE_PATH,
        "DATAVIEW_TABLE_TEMPLATE",
        default_dataview_table_template(),
    )
    return render_template(template, {"rows": js_rows}).splitlines()


def default_dataview_table_template() -> str:
    return """```dataviewjs
const rows = [{{rows}}];
dv.table(
  ["文献", "年份", "作者", "期刊", "DOI"],
  rows.map(row => [dv.parse(row.link), row.year, row.authors, row.journal, row.doi])
);
```"""


def escape_inline_field(value: str) -> str:
    return value.replace("\n", " ").replace("::", ":")


def escape_wikilink_part(value: str) -> str:
    return value.replace("|", "¦").replace("[", "(").replace("]", ")").replace("\n", " ")


def js_string(value: str) -> str:
    escaped = (
        value.replace("\\", "\\\\")
        .replace('"', '\\"')
        .replace("\n", "\\n")
        .replace("\r", "")
    )
    return f'"{escaped}"'


def truncate(value: str, max_length: int) -> str:
    value = clean_spaces(value)
    if len(value) <= max_length:
        return value
    return value[: max_length - 1].rstrip() + "…"


if __name__ == "__main__":
    main()
