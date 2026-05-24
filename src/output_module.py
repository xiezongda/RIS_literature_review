from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from AI_cache_annotation import build_analysis
from RIS_analysis import LiteratureRecord
from read_templete import (
    get_card_template,
    get_dataview_table_template,
    get_summary_header_template,
    render_template,
)


BASE_EXCEL_HEADERS = ["index", "title", "authors", "year", "journal", "doi", "keywords", "abstract"]
TAIL_EXCEL_HEADERS = ["source_files", "duplicate_count"]


def write_excel(
    records: list[LiteratureRecord],
    path: Path,
    annotations: dict[int, dict[str, str]],
    fields: list[str],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "literature_records"

    headers = build_excel_headers(fields)
    sheet.append(headers)

    for index, record in enumerate(records, start=1):
        context = build_record_context(index, record, fields, annotations.get(index, {}))
        sheet.append([context.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = default_column_width(header)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(records) + 1)}"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_excel_headers(fields: list[str]) -> list[str]:
    headers: list[str] = []
    for header in [*BASE_EXCEL_HEADERS, *fields, *TAIL_EXCEL_HEADERS]:
        if header not in headers:
            headers.append(header)
    return headers


def default_column_width(header: str) -> int:
    if header in {"index", "year", "duplicate_count"}:
        return 10
    if header in {"title", "abstract", "review_sentence"}:
        return 55
    if header in {"abstract_summary", "core_result", "fcva_connection", "research_topic_connection"}:
        return 46
    return 28


def write_literature_cards(
    records: list[LiteratureRecord],
    path: Path,
    annotations: dict[int, dict[str, str]],
    fields: list[str],
) -> None:
    lines: list[str] = [
        "# 文献卡片",
        "",
        f"> 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
    ]

    if not records:
        lines.extend(["当前未在 `data/raw/` 中读取到 RIS 文献。把 `.ris` 文件放入该目录后，重新运行 `python src/main.py`。", ""])
    else:
        for index, record in enumerate(records, start=1):
            lines.extend(render_record_card(index, record, fields, annotations.get(index, {})))
            lines.append("")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def render_record_card(index: int, record: LiteratureRecord, fields: list[str], annotation: dict[str, str]) -> list[str]:
    context = build_record_context(index, record, fields, annotation)
    return (render_template(get_card_template(), context).rstrip() + "\n").splitlines()


def build_record_context(index: int, record: LiteratureRecord, fields: list[str], annotation: dict[str, str]) -> dict[str, Any]:
    analysis = build_analysis(record, fields, annotation)
    title = record.title or "未命名文献"
    context: dict[str, Any] = {
        "index": index,
        "literature_id": f"L{index:03d}",
        "title": title,
        "title_inline": escape_inline_field(title),
        "authors": format_authors(record.authors),
        "year": record.year or "未提供",
        "journal": record.journal or "未提供",
        "doi": record.doi or "未提供",
        "keywords": "；".join(record.keywords) if record.keywords else "未提供",
        "abstract": record.abstract or "未提供",
        "source_files": "；".join(record.source_files),
        "duplicate_count": record.duplicate_count,
        **analysis,
    }
    return context


def write_summary(
    records: list[LiteratureRecord],
    path: Path,
    annotations: dict[int, dict[str, str]],
    fields: list[str],
) -> None:
    header = render_template(
        get_summary_header_template(),
        {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "record_count": len(records),
            "cards_file": "literature_cards.md",
        },
    )
    lines: list[str] = header.rstrip().splitlines()
    lines.append("")

    if not records:
        lines.extend(["当前没有可分类文献。把 `.ris` 文件放入 `data/raw/` 后运行 `python src/main.py`。", ""])
    else:
        grouped: dict[str, dict[str, list[tuple[int, LiteratureRecord]]]] = defaultdict(lambda: defaultdict(list))
        for index, record in enumerate(records, start=1):
            analysis = build_analysis(record, fields, annotations.get(index, {}))
            broad = analysis.get("broad_direction") or "未分类"
            medium = analysis.get("medium_direction") or "待人工分类"
            grouped[broad][medium].append((index, record))

        for broad in sorted(grouped):
            lines.extend([f"## {broad}", ""])
            for medium in sorted(grouped[broad]):
                lines.extend([f"### {medium}", ""])
                lines.extend(render_dataview_table(grouped[broad][medium]))
                lines.append("")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def print_pipeline_summary(
    ris_files: list[Path],
    raw_records: list[LiteratureRecord],
    records: list[LiteratureRecord],
    annotations: dict[int, dict[str, str]],
    fields: list[str],
    excel_path: Path,
    cards_path: Path,
    summary_path: Path,
) -> None:
    print(f"RIS 文件数量：{len(ris_files)}")
    print(f"原始文献条目：{len(raw_records)}")
    print(f"去重后条目：{len(records)}")
    print(f"AI/缓存标注条目：{len(annotations)}")
    print(f"动态 AI 输出字段：{', '.join(fields) if fields else '无'}")
    print(f"Excel 输出：{excel_path}")
    print(f"文献卡片：{cards_path}")
    print(f"分类总结：{summary_path}")


def render_dataview_table(items: list[tuple[int, LiteratureRecord]]) -> list[str]:
    rows = []
    for index, record in items:
        title = record.title or "未命名文献"
        heading = f"{index}. {title}"
        wikilink = f"[[literature_cards#{escape_wikilink_part(heading)}|{escape_wikilink_part(title)}]]"
        rows.append(
            {
                "link": wikilink,
                "year": record.year or "",
                "authors": format_authors(record.authors, limit=3),
                "journal": record.journal or "",
                "doi": record.doi or "",
            }
        )

    js_rows = ",\n  ".join(
        "{"
        f"link: {js_string(row['link'])}, "
        f"year: {js_string(row['year'])}, "
        f"authors: {js_string(row['authors'])}, "
        f"journal: {js_string(row['journal'])}, "
        f"doi: {js_string(row['doi'])}"
        "}"
        for row in rows
    )
    return render_template(get_dataview_table_template(), {"rows": js_rows}).splitlines()


def format_authors(authors: list[str], limit: int = 8) -> str:
    if not authors:
        return "未提供"
    shown = authors[:limit]
    suffix = " 等" if len(authors) > limit else ""
    return "；".join(shown) + suffix


def escape_inline_field(value: str) -> str:
    return value.replace("\n", " ").replace("::", ":")


def escape_wikilink_part(value: str) -> str:
    return value.replace("|", "¦").replace("[", "(").replace("]", ")").replace("\n", " ")


def js_string(value: str) -> str:
    escaped = value.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n").replace("\r", "")
    return f'"{escaped}"'
